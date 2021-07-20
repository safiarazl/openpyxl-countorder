import datetime, os
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("Book1.xlsx")
ws = wb.active

def nambahData(co,paid):
    a, b = ws['c2'], ws['c3']
    ws['c2'] = str(int(a.value) + int(co))
    ws['c3'] = str(int(b.value) + int(paid))
    ws['c6'] = 'Updated'
    ws['d6'] = datetime.datetime.now().strftime('%H:%M:%S - %d %B/%m %Y')
    wb.save("Book1.xlsx")

def ngurangData(co, paid):
    a, b = ws['c2'], ws['c3']
    ws['c2'] = str(int(a.value) - int(co))
    ws['c3'] = str(int(b.value) - int(paid))
    ws['c6'] = 'Updated'
    ws['d6'] = datetime.datetime.now().strftime('%H:%M:%S - %d %B/%m %Y')
    wb.save("Book1.xlsx")

def tampilkanData():
    print(f"{'='*3}Tampilan Data{'='*3}")
    print(f"Jumlah Checkout: {ws['c2'].value}\nJumlah Payment: {ws['c3'].value}\nLast Updated: {ws['d6'].value}")
    print(f"Selisih dari pendapatan bulan lalu: {ws.cell(5, 4)}")
    print('='*20)

if __name__ == "__main__":
    menu = ['Nambah Data', 'Mengurangi Data', 'Tampilkan Data','Exit']
    c = ws['c2']
    while True:
        for idx, item in enumerate(menu):
            print(f'{idx+1}. {item}')
        pilmen = int(input('option: '))
        if pilmen == 1:
            co = input("Checkout: ")
            paid = input("Paid: ")
            nambahData(co, paid)
        elif pilmen == 2:
            co = input("Checkout: ")
            paid = input("Paid: ")
            ngurangData(co, paid)
        elif pilmen == 3:
            os.system('clear')
            tampilkanData()
        elif pilmen == 4:
            break